#!/usr/bin/env python3
"""
Fetch NanoGPT available models from API

Usage:
    python fetch_nanogpt_models.py

Requires:
    - NANOGPT_API_KEY environment variable or .env file

Outputs:
    - nanogpt_models.csv
    - nanogpt_models.xlsx
    - nanogpt_models.json (raw API response)
"""

import json
import csv
import os
from pathlib import Path
from typing import List, Dict, Any

try:
    import requests
    HAS_REQUESTS = True
except ImportError:
    HAS_REQUESTS = False
    print("Error: requests not installed. Install with: pip install requests")
    exit(1)

try:
    import openpyxl
    from openpyxl.styles import Font
    HAS_OPENPYXL = True
except ImportError:
    HAS_OPENPYXL = False
    print("Warning: openpyxl not installed. XLSX output will be skipped.")
    print("Install with: pip install openpyxl")

try:
    from dotenv import load_dotenv
    load_dotenv()
except ImportError:
    pass  # dotenv optional


def fetch_models(api_key: str, base_url: str = "https://nano-gpt.com/api/v1") -> List[Dict[str, Any]]:
    """Fetch available models from NanoGPT API"""

    headers = {
        "Authorization": f"Bearer {api_key}",
        "Content-Type": "application/json"
    }

    url = f"{base_url}/models"

    print(f"Fetching models from {url}...")
    response = requests.get(url, headers=headers)
    response.raise_for_status()

    data = response.json()

    # API should return {"data": [...models...]}
    if isinstance(data, dict) and 'data' in data:
        models = data['data']
    elif isinstance(data, list):
        models = data
    else:
        raise ValueError(f"Unexpected API response format: {type(data)}")

    print(f"✓ Fetched {len(models)} models")

    return models


def save_json(models: List[Dict], output_path: Path):
    """Save raw JSON response"""
    with open(output_path, 'w', encoding='utf-8') as f:
        json.dump(models, f, indent=2)
    print(f"✓ Saved JSON to {output_path}")


def models_to_rows(models: List[Dict]) -> tuple[List[str], List[List[str]]]:
    """Convert models list to CSV rows"""

    # Determine all available fields
    all_fields = set()
    for model in models:
        all_fields.update(model.keys())

    # Common fields to prioritize in column order
    priority_fields = ['id', 'name', 'created', 'owned_by', 'object']

    # Build headers: priority fields first, then alphabetically sorted rest
    headers = []
    for field in priority_fields:
        if field in all_fields:
            headers.append(field)
            all_fields.remove(field)

    # Add remaining fields alphabetically
    headers.extend(sorted(all_fields))

    # Build rows
    rows = []
    for model in models:
        row = []
        for field in headers:
            value = model.get(field, '')
            # Convert non-string values to strings
            if isinstance(value, (list, dict)):
                value = json.dumps(value)
            else:
                value = str(value) if value is not None else ''
            row.append(value)
        rows.append(row)

    return headers, rows


def save_csv(headers: List[str], rows: List[List[str]], output_path: Path):
    """Save to CSV"""
    with open(output_path, 'w', newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        writer.writerow(headers)
        writer.writerows(rows)
    print(f"✓ Saved CSV to {output_path}")


def save_xlsx(headers: List[str], rows: List[List[str]], output_path: Path):
    """Save to XLSX with formatting"""
    if not HAS_OPENPYXL:
        print("✗ Skipping XLSX (openpyxl not installed)")
        return

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "NanoGPT Models"

    # Write headers
    for col_idx, header in enumerate(headers, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = Font(bold=True)

    # Write data
    for row_idx, row in enumerate(rows, start=2):
        for col_idx, value in enumerate(row, start=1):
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Auto-adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if cell.value and len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(max_length + 2, 60)
        ws.column_dimensions[column_letter].width = adjusted_width

    wb.save(output_path)
    print(f"✓ Saved XLSX to {output_path}")


def main():
    # Get API key
    api_key = os.getenv('NANOGPT_API_KEY')
    if not api_key:
        print("Error: NANOGPT_API_KEY not found in environment")
        print("Set it with: export NANOGPT_API_KEY='your-key-here'")
        print("Or create a .env file with: NANOGPT_API_KEY=your-key-here")
        return 1

    # Paths
    project_root = Path(__file__).parent.parent
    json_path = project_root / "nanogpt_models.json"
    csv_path = project_root / "nanogpt_models.csv"
    xlsx_path = project_root / "nanogpt_models.xlsx"

    try:
        # Fetch models
        models = fetch_models(api_key)

        # Save raw JSON
        save_json(models, json_path)

        # Convert to table format
        headers, rows = models_to_rows(models)

        # Save outputs
        save_csv(headers, rows, csv_path)
        save_xlsx(headers, rows, xlsx_path)

        # Print summary
        print(f"\n📊 Summary:")
        print(f"   Models: {len(models)}")
        print(f"   Fields: {len(headers)}")
        print(f"   Columns: {', '.join(headers[:5])}...")

        # Show a few examples
        if models:
            print(f"\n📝 Sample models:")
            for model in models[:5]:
                model_id = model.get('id', 'unknown')
                model_name = model.get('name', model_id)
                print(f"   - {model_name} ({model_id})")

        return 0

    except Exception as e:
        print(f"Error: {e}")
        import traceback
        traceback.print_exc()
        return 1


if __name__ == '__main__':
    exit(main())
